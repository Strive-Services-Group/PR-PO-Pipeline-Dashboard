#!/usr/bin/env python3
"""
Regenerate pr_steps.json (the authoritative Step-name lookup) from a fresh D365 PR export.

The dashboard's Step name / step date / approver / department-location-contract come from
custom IFAHR* fields stored on PurchReqTable. Those aren't on a live OData entity yet, so the
fresh export is the source of truth. This script turns the export into pr_steps.json, which
the dashboard overlays onto the live proxy data by requisition number.

Usage:
    python gen_pr_steps.py pr.xlsx
Then commit pr_steps.json and push.
"""
import sys, json, datetime
import openpyxl

src = sys.argv[1] if len(sys.argv) > 1 else "pr.xlsx"
wb = openpyxl.load_workbook(src, read_only=True)
ws = wb.active
hdr = [c for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]

def col(name):
    return hdr.index(name) if name in hdr else None

ix = {k: col(k) for k in [
    'Purchase requisition', 'Step name', 'Step date and time', 'Pending Approver/User',
    'Accepted By/Assign To', 'Department', 'Location', 'Contract',
    'Submission Status', 'Request for quotation case']}

def iso(v):
    return v.isoformat() if isinstance(v, (datetime.datetime, datetime.date)) else v

out = {}
for r in ws.iter_rows(min_row=2, values_only=True):
    pr = r[ix['Purchase requisition']]
    if not pr:
        continue
    out[pr] = {
        'step': r[ix['Step name']],
        'stepDate': iso(r[ix['Step date and time']]),
        'pendingUser': r[ix['Pending Approver/User']],
        'acceptedBy': r[ix['Accepted By/Assign To']],
        'department': r[ix['Department']],
        'location': r[ix['Location']],
        'contract': r[ix['Contract']],
        'submissionStatus': r[ix['Submission Status']],
        'rfqCase': r[ix['Request for quotation case']],
    }

with open("pr_steps.json", "w") as f:
    json.dump(out, f, separators=(',', ':'))

print(f"Wrote pr_steps.json: {len(out)} requisitions, "
      f"{sum(1 for v in out.values() if v['step'])} with an active step.")
